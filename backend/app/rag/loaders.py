"""Document loader factory supporting multiple file types."""
from pathlib import Path
from langchain_core.documents import Document


def load_document(file_path: str) -> list[Document]:
    """Load a document using the appropriate LangChain loader."""
    ext = Path(file_path).suffix.lower()

    if ext == ".pdf":
        from langchain_community.document_loaders import PyPDFLoader
        loader = PyPDFLoader(file_path)
    elif ext == ".docx":
        from langchain_community.document_loaders import Docx2txtLoader
        loader = Docx2txtLoader(file_path)
    elif ext in (".txt", ".md"):
        from langchain_community.document_loaders import TextLoader
        loader = TextLoader(file_path, encoding="utf-8")
    elif ext == ".csv":
        from langchain_community.document_loaders import CSVLoader
        loader = CSVLoader(file_path, encoding="utf-8")
    elif ext == ".xlsx":
        return _load_xlsx(file_path)
    else:
        raise ValueError(f"Unsupported file type: {ext}")

    return loader.load()


def _load_xlsx(file_path: str) -> list[Document]:
    """Load Excel file row-by-row as documents."""
    import openpyxl
    from langchain_core.documents import Document

    wb = openpyxl.load_workbook(file_path, read_only=True)
    docs = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
        for row_idx, row in enumerate(rows[1:], start=2):
            content_parts = [f"{headers[i]}: {val}" for i, val in enumerate(row) if val is not None]
            if content_parts:
                content = "\n".join(content_parts)
                docs.append(
                    Document(
                        page_content=content,
                        metadata={
                            "source": file_path,
                            "sheet": sheet_name,
                            "row": row_idx,
                        },
                    )
                )
    wb.close()
    return docs
